import os
import re
import wave

import pandas as pd
import torch
import faster_whisper
import numpy as np
from openpyxl import load_workbook

MODEL_DIRECTORY = "models"
MODEL_NAME = "faster-whisper-small-30185-1385-18168"

BASE_AUDIO_FOLDER = "chunked_audio"
SPECIFIC_AUDIO_FOLDER = "CYOW-App-Dep-May-10-2026-1200Z-2000Z"  # Change this to your actual folder path

MODEL = os.path.join(MODEL_DIRECTORY, MODEL_NAME)
INPUT_FULL_PATH = os.path.join(BASE_AUDIO_FOLDER, SPECIFIC_AUDIO_FOLDER)

# configure model
batch_size = 8

language = None  # autodetect language

device = "cuda" if torch.cuda.is_available() else "cpu"
print(device)

# compute_type = "float16"
compute_type = "int8"
if device == "cuda":
    compute_type = "float16"

print(compute_type)

whisper_model = faster_whisper.WhisperModel(
    MODEL, device=device, compute_type=compute_type
)
whisper_pipeline = faster_whisper.BatchedInferencePipeline(whisper_model)

# extract the trailing number before .wav
def extract_trailing_number(filename):
    match = re.search(r'_(\d+)\.wav$', filename)
    return int(match.group(1)) if match else float('inf')

def find_numeral_tokens(tokenizer):
    return [
        i
        for i in range(tokenizer.eot)
        if (all(c in "0123456789" for c in tokenizer.decode([i]).removeprefix(" "))
            and len(tokenizer.decode([i]).strip()) > 0)
    ]

# turns generator object to a JSON like structure
def jsonify_segments(segment):
    return {
        "id": segment.id,
        "start": segment.start,
        "end": segment.end,
        "text": segment.text,
        "words": [
            {
                "start": w.start,
                "end": w.end,
                "word": w.word,
                "probability": w.probability
            }
            for w in segment.words # word is a list
        ],
        "no_speech_prob": segment.no_speech_prob
    }

def get_transcription(audio_file):
    audio_waveform = faster_whisper.decode_audio(audio_file)

    # Preprocess waveform before transcription
    # audio_waveform = bandpass_filter(audio_waveform)
    # audio_waveform = normalize_audio(audio_waveform)
    audio_waveform = audio_waveform.astype(np.float32)

    transcript_segments, info = whisper_pipeline.transcribe(
        audio_waveform,
        language=language,
        initial_prompt="",
        suppress_tokens=[-1] + find_numeral_tokens(faster_whisper.tokenizer.Tokenizer(tokenizer=whisper_model.hf_tokenizer, multilingual=False)),
        batch_size=batch_size,
        word_timestamps=True,
        task="transcribe",
    )

    torch.cuda.empty_cache()

    return transcript_segments

def transcribe_audio(audio_file):
    try:
        # no need to chunk since imported audio should already be in chunked form
        transcript = get_transcription(audio_file)
        segments_dict = [jsonify_segments(s) for s in list(transcript)]

        total_probability = 0
        count = 0

        for segment in segments_dict:
            for word in segment["words"]:
                total_probability += word["probability"]
                count += 1

        average_probability = total_probability / count if count > 0 else 0

        # ignore metadata and only get the full text, probability metadata may be useful for WER
        text = " ".join([entry["text"] for entry in segments_dict])
        return text, average_probability
    except Exception as e:
        print(f"[ASR error] {os.path.basename(audio_file)} -> {e}")
        return None, None

# Sort files by extracted number
file_list = [f for f in os.listdir(INPUT_FULL_PATH) if f.endswith(".wav") and "_" in f]
sorted_files = sorted(file_list, key=extract_trailing_number)

# Prepare a list to collect results
results = []

# Iterate through files in sorted order
for file_name in sorted_files:
    if file_name.endswith(".wav"):
        wav_path = os.path.join(INPUT_FULL_PATH, file_name)
        print(f"Processing: {wav_path}")

        # Calculate duration in milliseconds and seconds
        with wave.open(wav_path, 'r') as wav_file:
            frames = wav_file.getnframes()
            rate = wav_file.getframerate()
            duration_sec = frames / float(rate)
            duration_ms = duration_sec * 1000

        # Transcribe the audio
        transcript, _ = transcribe_audio(wav_path)

        # Append the result
        results.append({
            "file_name": file_name,
            "Clarity": "",
            "transcript": transcript,
            "Comment": "",
            "duration": round(duration_sec, 4)
        })

# Create a DataFrame and save to Excel
df = pd.DataFrame(results)
df.to_excel(f"{INPUT_FULL_PATH}.xlsx", index=False)

wb = load_workbook(f"{INPUT_FULL_PATH}.xlsx")
ws = wb.active  # first sheet

# Set custom values
ws["F1"] = "1=Unusable, 2=Mostly Unusable, 3=Partly Unusable/Heavy Noise, 4=Moderate Noise, 5=Minor Noise, 6=No Noise"
ws["F2"] = '=SUMIFS(E2:E1048576, B2:B1048576, ">=4")'  # formula

# 5️⃣ Save workbook
wb.save(f"{INPUT_FULL_PATH}.xlsx")

print(f"Transcription complete. Results saved to {INPUT_FULL_PATH}.xlsx")